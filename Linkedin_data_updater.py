# Import libraries.
# import numpy as np
import pandas as pd

print('Remember to backup your data Linkedin_xxxxx.xlsx files before you continue')
print('Be sure the files you downloaded are on the same directory as this executable file')

def confirmation_dialog(confirm):
    while confirm != 'N':
        confirm = input('Update xlsx? y=Continue / n=exit program: ')
        if confirm == 'y':
            break
        elif confirm == 'n':
            exit()
        elif confirm == 'Y':
            break
        elif confirm == 'N':
            exit()

# FOLLOWERS DATA UPDATER

print('FOLLOWERS DATA')
print()

# Input the downloaded file name from Linkedin containing the new Followers info
while True:
    print('Enter file name of downloaded Linkedin followers data .xls or .xlsx')

    file_name_followers = input()

    if (file_name_followers.find('followers') != -1):
        print ("File name correct")
        print('File name to be updated:', file_name_followers)
        break
    else:
        print ("File name incorrect. Please check if the file contains Visitors data")
    continue

# Just for checking I list the names of the sheets inside the file.
followers = pd.ExcelFile(file_name_followers)
shs = followers.sheet_names
print('The following sheets will be updated:', shs)

# Load each of the sheets in the update file to a different Pandas dataframe.
# New followers sheet is different than the rest in the data structure.
new_followers = pd.read_excel(file_name_followers,
                              sheet_name='New followers')

# Location, Job function, Seniority, Industry and Company size sheets have equal structure.
followers_Location = pd.read_excel(file_name_followers,
                                   sheet_name='Location')

followers_Job_function = pd.read_excel(file_name_followers,
                                       sheet_name='Job function')

followers_Seniority = pd.read_excel(file_name_followers,
                                    sheet_name='Seniority')

followers_Industry = pd.read_excel(file_name_followers,
                                   sheet_name='Industry')

followers_Company_size = pd.read_excel(file_name_followers,
                                       sheet_name='Company size')

# Convert new_followers Date field to datetime.
new_followers['Date'] = pd.to_datetime(new_followers['Date'])

# I define the date of the files download. This is the last day in which the files have been downloaded from Linkedin.
download_date = new_followers['Date'].max()
download_date = pd.to_datetime(download_date, format='%Y-%m-%d')
print('Last day downloaded from Linkedin:', download_date)

# Add download_date to each of the detailed followers dataframes. This will give us a date for the amount and distribution of each of the sheets in the original downloaded file.
followers_Location['Date'] = download_date
followers_Job_function['Date'] = download_date
followers_Seniority['Date'] = download_date
followers_Industry['Date'] = download_date
followers_Company_size['Date'] = download_date

# Read the Linkedin_followers_New_followers.csv file in which the new data will be stored
Linkedin_followers = pd.read_excel('Linkedin_followers.xlsx')

# Define last_date as the last available date in Linkedin_followers_New_followers.csv
last_date = Linkedin_followers['Date'].max()
if pd.isna(last_date):
    last_date = '2001-01-01'
    print('No data on database. New info will be updated')
else:
    print('Previous last date available in Linkedin_followers.xlsx was:',  last_date)
print('Data will be updated up to:', download_date)
print()

confirm = ''
confirmation_dialog(confirm)

# new_followers dataframe is trimmed to include only the data that is not on Linkedin_followers_New_followers.csv. Data from last_date to download_date (included).
new_followers = new_followers.loc[(new_followers['Date'] > last_date) & (new_followers['Date'] <= download_date)]

# Write new info on the excel master data file
from openpyxl import load_workbook

with pd.ExcelWriter('Linkedin_followers.xlsx', mode='a',datetime_format='yyyy-mm-dd', engine="openpyxl") as writer:
    writer.book = load_workbook('Linkedin_followers.xlsx')
    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)

    new_followers.to_excel(writer, sheet_name='New followers', index=False, header=False,
                           startrow=len(pd.read_excel('Linkedin_followers.xlsx', sheet_name='New followers'))+1)
    followers_Location.to_excel(writer, sheet_name='Location', index=False, header=False,
                                startrow=len(pd.read_excel('Linkedin_followers.xlsx', sheet_name='Location'))+1)
    followers_Job_function.to_excel(writer, sheet_name='Job function', index=False, header=False,
                                    startrow=len(pd.read_excel('Linkedin_followers.xlsx', sheet_name='Job function'))+1)
    followers_Seniority.to_excel(writer, sheet_name='Seniority', index=False, header=False,
                                 startrow=len(pd.read_excel('Linkedin_followers.xlsx', sheet_name='Seniority'))+1)
    followers_Industry.to_excel(writer, sheet_name='Industry', index=False, header=False,
                                startrow=len(pd.read_excel('Linkedin_followers.xlsx', sheet_name='Industry'))+1)
    followers_Company_size.to_excel(writer, sheet_name='Company size', index=False, header=False,
                                    startrow=len(pd.read_excel('Linkedin_followers.xlsx', sheet_name='Company size'))+1)

writer.save()

print()
print('FINISHED UPDATING FOLLOWERS DATA.')
print('---------------')
print()
print()
# VISITORS DATA UPDATER

print('VISITORS DATA')
print()

# Input the downloaded file name from Linkedin containing the new Visitors info
while True:
    print('Enter file name of downloaded Linkedin visitors data .xls or .xlsx')

    file_name_visitors = input()

    if (file_name_visitors.find('visitors') != -1):
        print ("File name correct")
        print('File name to be updated:', file_name_visitors)
        break
    else:
        print ("File name incorrect. Please check if the file contains Visitors data")
    continue

# Just for checking I list the names of the sheets inside the file.
visitors = pd.ExcelFile(file_name_visitors)
shs = visitors.sheet_names
print('The following sheets will be updated:', shs)

# Load each of the sheets in the update file to a different Pandas dataframe.
# New visitors sheet is different than the rest in the data structure.
visitor_metrics = pd.read_excel(file_name_visitors,
                              sheet_name='Visitor metrics')

# Location, Job function, Seniority, Industry and Company size sheets have equal structure.
visitors_Location = pd.read_excel(file_name_visitors,
                                   sheet_name='Location')

visitors_Job_function = pd.read_excel(file_name_visitors,
                                       sheet_name='Job function')

visitors_Seniority = pd.read_excel(file_name_visitors,
                                    sheet_name='Seniority')

visitors_Industry = pd.read_excel(file_name_visitors,
                                   sheet_name='Industry')

visitors_Company_size = pd.read_excel(file_name_visitors,
                                       sheet_name='Company size')

# Convert new_visitors Date field to datetime.
visitor_metrics['Date'] = pd.to_datetime(visitor_metrics['Date'])

# I define the date of the files download. This is the last day in which the files have been downloaded from Linkedin.
download_date = visitor_metrics['Date'].max()
download_date = pd.to_datetime(download_date, format='%Y-%m-%d')
print('Last day downloaded from Linkedin:', download_date)

# Add download_date to each of the detailed visitors dataframes. This will give us a date for the amount and distribution of each of the sheets in the original downloaded file.
visitors_Location['Date'] = download_date
visitors_Job_function['Date'] = download_date
visitors_Seniority['Date'] = download_date
visitors_Industry['Date'] = download_date
visitors_Company_size['Date'] = download_date

# Read the Linkedin_visitors_New_visitors.csv file in which the new data will be stored
Linkedin_visitors = pd.read_excel('Linkedin_visitors.xlsx')

# Define last_date as the last available date in Linkedin_visitors.xls
last_date = Linkedin_visitors['Date'].max()
if pd.isna(last_date):
    last_date = '2001-01-01'
    print('No data on database. New info will be updated')
else:
    print('Previous last date available in Linkedin_visitors.xlsx was:',  last_date)
print('Data will be updated up to:', download_date)
print()

# visitors_metrics dataframe is trimmed to include only the data that is not on Linkedin_followers_New_followers.csv. Data from last_date to download_date (included).
visitor_metrics = visitor_metrics.loc[(visitor_metrics['Date'] > last_date) & (visitor_metrics['Date'] <= download_date)]

confirm = ''
confirmation_dialog(confirm)

# Write new data on master excel file
with pd.ExcelWriter('Linkedin_visitors.xlsx',
                    mode='a',
                    datetime_format='yyyy-mm-dd',
                    engine="openpyxl") as writer:
    writer.book = load_workbook('Linkedin_visitors.xlsx')
    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)

    visitor_metrics.to_excel(writer, sheet_name='Visitor metrics', index=False, header=False,
                           startrow=len(pd.read_excel('Linkedin_visitors.xlsx', sheet_name='Visitor metrics'))+1)
    visitors_Location.to_excel(writer, sheet_name='Location', index=False, header=False,
                                startrow=len(pd.read_excel('Linkedin_visitors.xlsx', sheet_name='Location'))+1)
    visitors_Job_function.to_excel(writer, sheet_name='Job function', index=False, header=False,
                                    startrow=len(pd.read_excel('Linkedin_visitors.xlsx', sheet_name='Job function'))+1)
    visitors_Seniority.to_excel(writer, sheet_name='Seniority', index=False, header=False,
                                 startrow=len(pd.read_excel('Linkedin_visitors.xlsx', sheet_name='Seniority'))+1)
    visitors_Industry.to_excel(writer, sheet_name='Industry', index=False, header=False,
                                startrow=len(pd.read_excel('Linkedin_visitors.xlsx', sheet_name='Industry'))+1)
    visitors_Company_size.to_excel(writer, sheet_name='Company size', index=False, header=False,
                                    startrow=len(pd.read_excel('Linkedin_visitors.xlsx', sheet_name='Company size'))+1)

writer.save()

print()
print('FINISHED UPDATING VISITORS DATA.')
print('---------------')
print()
print()
# UPDATES DATA UPDATER

print('UPDATES DATA')
print()

# Input the downloaded file name from Linkedin containing the new Updates info
while True:
    print('Enter file name of downloaded Linkedin UPDATES data .xls or .xlsx')

    file_name_updates = input()

    if (file_name_updates.find('updates') != -1):
        print ("File name correct")
        print('File name to be updated:', file_name_updates)
        break
    else:
        print ("File name incorrect. Please check if the file contains Updates data")
    continue

# Just for checking I list the names of the sheets inside the file.
updates = pd.ExcelFile(file_name_updates)
shs = updates.sheet_names
print('The following sheets will be updated:', shs)

# Load each of the sheets in the udate file to a different Pandas dataframe.
new_metrics = pd.read_excel(file_name_updates,
                               sheet_name='Update metrics (aggregated)',
                               header=1
                              )

new_engagement = pd.read_excel(file_name_updates,
                               sheet_name='Update engagement',
                               header=1
                              )

# Convert update_metrics Date field to datetime.
new_metrics['Date'] = pd.to_datetime(new_metrics['Date'])
# Convert update_engagement dates fields to datetime.
new_engagement['Created date'] = pd.to_datetime(new_engagement['Created date'])
new_engagement['Campaign start date'] = pd.to_datetime(new_engagement['Campaign start date'])
new_engagement['Campaign end date'] = pd.to_datetime(new_engagement['Campaign end date'])

# Define download_date as the last date with data in the downloaded file from Linkedin
first_date = new_metrics['Date'].min()
download_date = new_metrics['Date'].max()

# Load master file Linkedin_updates.xls on Linkedin_updates_metrics dataframe
old_metrics = pd.read_excel('Linkedin_updates.xlsx',sheet_name='Update metrics (aggregated)')
old_engagement = pd.read_excel('Linkedin_updates.xlsx', sheet_name='Update engagement')

# Convert update_metrics Date field to datetime.
old_metrics['Date'] = pd.to_datetime(old_metrics['Date'])

# Convert update_engagement dates fields to datetime.
old_engagement['Created date'] = pd.to_datetime(old_engagement['Created date'])
old_engagement['Campaign start date'] = pd.to_datetime(old_engagement['Campaign start date'])
old_engagement['Campaign end date'] = pd.to_datetime(old_engagement['Campaign end date'])

# Define last_date as the last date available in master update file
last_date = old_metrics['Date'].max()
if pd.isna(last_date):
    last_date = '2001-01-01'
    print('No data on database. New info will be updated')
else:
    print('Previous last date available in Linkedin_updates.xlsx was:',  last_date)
print('Data will be updated up to:', download_date)
print()

old_engagement.set_index('Update link', inplace=True)

new_engagement['Type of content'] = np.nan
new_engagement['Content Format'] = np.nan
new_engagement['Language'] = np.nan
new_engagement.set_index('Update link', inplace=True)

df_engagement = old_engagement.append(new_engagement)
df_engagement['Type of content'].update(old_engagement['Type of content'])
df_engagement['Content Format'].update(old_engagement['Content Format'])
df_engagement['Language'].update(old_engagement['Language'])

df_engagement=df_engagement.drop_duplicates(subset=['Update title'], keep='last')

#  Trim the data from the downloaded file to update the info from the last date available on the master file.
df_metrics = old_metrics.loc[(old_metrics['Date'] < first_date)]

df_metrics = df_metrics.append(new_metrics, ignore_index=True)

confirm = ''
confirmation_dialog(confirm)

with pd.ExcelWriter('Linkedin_updates.xlsx',
#                     mode='a',
                    datetime_format='yyyy-mm-dd',
                    engine="openpyxl") as writer:
    writer.book = load_workbook('Linkedin_updates.xlsx')
    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)

    df_metrics.to_excel(writer, sheet_name='Update metrics (aggregated)', index=False, header=True)

    df_engagement.to_excel(writer, sheet_name='Update engagement', index=True, header=True)


writer.save()

print()
print('FINISHED UPDATING UPDATES DATA.')
print('---------------------')
print('Program ended. Check updated files.')
print('---------------------')
